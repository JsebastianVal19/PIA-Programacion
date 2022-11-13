# Estructura de datos y su procesamiento

# PIA

# Renta de espacios de coworking

# Imports

import datetime
import openpyxl
import sqlite3
from sqlite3 import Error
import sys

# Variables

dia = 1
dia_dos = 2
dia_tres = 3
fecha_hoy = datetime.date.today()

# Consultar datos clientes

try:
    with sqlite3.connect("Evidencia.db") as conn:
        mi_cursor = conn.cursor()
        mi_cursor.execute("CREATE TABLE IF NOT EXISTS Clientes_registrados (Clave_cliente INTEGER PRIMARY KEY, Nombre_cliente TEXT NOT NULL);")

except Error as e:
    print (e)

except :
    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")

try:
    with sqlite3.connect("Evidencia.db") as conn:
        mi_cursor = conn.cursor()
        mi_cursor.execute("SELECT Nombre_cliente FROM Clientes_registrados")
        nombre_cliente= mi_cursor.fetchall()

except Error as e:
    print (e)

except:
    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")

finally:
    conn.close()

# Consultar datos Eventos

try:
    with sqlite3.connect("Evidencia.db") as conn:
        mi_cursor = conn.cursor()
        mi_cursor.execute("CREATE TABLE IF NOT EXISTS Eventos_registrados (Clave_evento INTEGER PRIMARY KEY, Nombre_cliente TEXT NOT NULL, Nombre_sala TEXT NOT NULL, Asistentes INTEGER, Fecha timestamp, Turno INTEGER);")

except Error as e:
    print (e)

except:
    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")

try:
    with sqlite3.connect("Evidencia.db") as conn:
        mi_cursor = conn.cursor()
        mi_cursor.execute("SELECT Nombre_sala FROM Eventos_registrados")
        nombre_sala= mi_cursor.fetchall()

except Error as e:
    print (e)

except:
    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")

finally:
    conn.close()

base_de_datos = nombre_cliente, nombre_sala

if base_de_datos :
    if base_de_datos[0] :
        pass

    elif base_de_datos[1] :
        pass

    else :
        print ("\n--------------------------------------------------------------")
        print ("No se encontraron datos en la base de datos de nuestro sistema")
        print ("Felicidades,eres el primer usuario en utilizar nuestro sistema")
        print ("--------------------------------------------------------------")

# Menu

while True :

    print ("\n****************")
    print ("   BIENVENIDO   ")
    print ("****************")

    print ("\n MENU PRINCIPAL")
    print ("\n[1] Revisar las reservaciones")
    print ("\n[2] Revisar los reportes")
    print ("\n[3] Registrar una sala")
    print ("\n[4] Registrase como nuevo cliente")
    print ("\n[5] Salir del programa")

    menu_principal = int (input ("\nPor favor ingresa el numero del menu al que deseas acceder : "))

    # Menu opcion 1

    if menu_principal == 1 :

        while True :
            print ("\n-- REVISAR LAS RESERVACIONES --")

            print ("\nMENU")
            print ("\n[1] Editar el nombre del evento de una reservacion ya hecha")
            print ("\n[2] Consulatar la disponibilidad de salas para una fecha")
            print ("\n[3] Eliminar una reservacion")
            print ("\n[4] Regresar al menu principal")

            sub_menu = int (input ("\nPor favor escribe el numero del menu al que desea acceder : "))

            # Editar el nombre de una reservacion

            if sub_menu == 1 :
                print ("\n-- EDITAR EL NOMBRE DEL EVENTO DE UNA RESERVACION YA HECHA --")

                try:
                    with sqlite3.connect("Evidencia.db") as conn:
                        mi_cursor = conn.cursor()
                        mi_cursor.execute("CREATE TABLE IF NOT EXISTS Eventos_registrados (Clave_evento INTEGER PRIMARY KEY, Nombre_cliente TEXT NOT NULL, Nombre_sala TEXT NOT NULL, Asistentes INTEGER, Fecha timestamp, Turno INTEGER);")

                except Error as e:
                    print (e)

                except:
                    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")

                acceso_dos = int (input ("\nPor favor ingresa la clave de tu evento : "))

                try:
                    with sqlite3.connect("Evidencia.db") as conn:
                        mi_cursor = conn.cursor()
                        valores = {"Clave_evento": acceso_dos}
                        mi_cursor.execute("SELECT Nombre_sala FROM Eventos_registrados WHERE Clave_evento = :Clave_evento", valores)
                        nombre_sala= mi_cursor.fetchall()

                        if nombre_sala :
                            print ("\n-----------------")
                            print ("Evento encontrado")
                            print ("-----------------")

                            print ("\nEl nombre de tu evento es : ", nombre_sala[0][0])

                            try:
                                with sqlite3.connect("Evidencia.db") as conn:
                                    mi_cursor = conn.cursor()

                                    nuevo_nombre = input ("\nEscribe el nuevo nombre para tu evento : ")

                                    valores = (nuevo_nombre, acceso_dos)

                                    mi_cursor.execute("UPDATE Eventos_registrados SET Nombre_sala=(?) WHERE Clave_evento=(?);",valores)

                            except Error as e:
                                print (e)

                            except:
                                print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")

                            finally:
                                conn.close()

                            print ("\n-------------------------------------------")
                            print ("El nombre del evento se edito correctamente")
                            print ("\nEl nuevo nombre de tu evento es : ", nuevo_nombre)
                            print ("-------------------------------------------")

                            regreso = input ('\nEscribe "0" para regresar al menu : ')

                        else :
                            print ("\n--------------------")
                            print ("Evento no encontrado")
                            print ("--------------------")

                            regreso = input ('\nEscribe "0" para regresar al menu : ')

                except Error as e:
                    print (e)

                except:
                    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")

                finally:
                    conn.close()

            # Disponibilidad de salas

            elif sub_menu == 2 :
                print ("\n-- CONSULTAR LA DISPONIBILIDAD DE SALAS PARA UNA FECHA --")

                print ('\nEscribe la fecha con el siguiente formato "09/01/2004"')
                consulta_dos = input("\nEscribe la fecha en la que quieres consultar las reservaciones disponibles : ")
                consulta_dos = datetime.datetime.strptime (consulta_dos,'%d/%m/%Y').date()

                # Base de datos eventos Fecha - Turno - Universal

                try:
                    with sqlite3.connect("Evidencia.db") as conn:
                        mi_cursor = conn.cursor()
                        mi_cursor.execute("CREATE TABLE IF NOT EXISTS Fecha_turno (Fecha timestamp, Turno INTEGER);")

                except Error as e:
                    print (e)

                except:
                    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")

                try:
                    with sqlite3.connect("Evidencia.db") as conn:
                        mi_cursor = conn.cursor()
                        valores = {"Fecha":consulta_dos,"Turno":1}
                        mi_cursor.execute("INSERT INTO Fecha_turno (Fecha, Turno) VALUES (:Fecha,:Turno)", valores)

                        valores = {"Fecha":consulta_dos,"Turno":2}
                        mi_cursor.execute("INSERT INTO Fecha_turno (Fecha, Turno) VALUES (:Fecha,:Turno)", valores)

                        valores = {"Fecha":consulta_dos,"Turno":3}
                        mi_cursor.execute("INSERT INTO Fecha_turno (Fecha, Turno) VALUES (:Fecha,:Turno)", valores)

                        valores = {"Fecha":consulta_dos}
                        mi_cursor.execute("SELECT Fecha, Turno FROM Fecha_turno WHERE Fecha = :Fecha", valores)
                        fechas_turnos_universales = mi_cursor.fetchall()
                        fechas_turnos_universales = set (fechas_turnos_universales)

                except Error as e:
                    print (e)

                except:
                    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")

                finally:
                    conn.close()

                # Seleccionar fechas y turnos de los eventos

                try:
                    with sqlite3.connect("Evidencia.db") as conn:
                        mi_cursor = conn.cursor()
                        mi_cursor.execute("CREATE TABLE IF NOT EXISTS Eventos_registrados (Clave_evento INTEGER PRIMARY KEY, Nombre_cliente TEXT NOT NULL, Nombre_sala TEXT NOT NULL, Asistentes INTEGER, Fecha timestamp, Turno INTEGER);")

                except Error as e:
                    print (e)

                except:
                    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")

                try:
                    with sqlite3.connect("Evidencia.db") as conn:
                        mi_cursor = conn.cursor()
                        valores = {"Fecha":consulta_dos}
                        mi_cursor.execute("SELECT Fecha, Turno FROM Eventos_registrados WHERE Fecha = :Fecha", valores)
                        fechas_turnos_ocupados = mi_cursor.fetchall()
                        fechas_turnos_ocupados = set (fechas_turnos_ocupados)

                except Error as e:
                    print (e)

                except:
                    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")

                finally:
                    conn.close()

                fechas_turnos_disponibles = fechas_turnos_universales - fechas_turnos_ocupados
                fechas_turnos_disponibles = list (fechas_turnos_disponibles)
                fechas_turnos_disponibles.sort()

                print ('\nLos turnos disponibles para el dia "', consulta_dos,'" son los siguientes')
                print ("\n------------------------------------------")
                print ("Numero de sala - Nombre de la sala - Turno")
                print ("------------------------------------------")

                for fecha,turno in fechas_turnos_disponibles :
                    print ("------------------------------------------")
                    print ("       1", "             Premium         " ,turno)
                    print ("------------------------------------------")

                regreso = input ('\nEscribe "0" para regresar al menu : ')

            # Eliminar una reservacion

            elif sub_menu == 3 :
                print ("\n-- ELIMINAR UNA RESERVACION --")

                try:
                    with sqlite3.connect("Evidencia.db") as conn:
                        mi_cursor = conn.cursor()
                        mi_cursor.execute("CREATE TABLE IF NOT EXISTS Eventos_registrados (Clave_evento INTEGER PRIMARY KEY, Nombre_cliente TEXT NOT NULL, Nombre_sala TEXT NOT NULL, Asistentes INTEGER, Fecha timestamp, Turno INTEGER);")

                except Error as e:
                    print (e)

                except:
                    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")

                print ("\nSolamente se puede eliminar una reservacion con tres dias de anticipacion")
                nuevo_acceso = int (input("\nPor favor ingresa la clave de tu evento : "))

                try:
                    with sqlite3.connect("Evidencia.db") as conn:
                        mi_cursor = conn.cursor()
                        valores = {"Clave_evento": nuevo_acceso}
                        mi_cursor.execute("SELECT Clave_evento, Nombre_sala, Asistentes, Fecha, Turno FROM Eventos_registrados WHERE Clave_evento = :Clave_evento", valores)
                        evento = mi_cursor.fetchall()

                        valores = {"Clave_evento": nuevo_acceso}
                        mi_cursor.execute("SELECT Clave_evento FROM Eventos_registrados WHERE Clave_evento = :Clave_evento", valores)
                        clave = mi_cursor.fetchall()

                        valores = {"Clave_evento": nuevo_acceso}
                        mi_cursor.execute("SELECT Fecha FROM Eventos_registrados WHERE Clave_evento = :Clave_evento", valores)
                        fecha = mi_cursor.fetchall()

                        if clave :
                            print ("\n-----------------")
                            print ("Evento encontrado")
                            print ("-----------------")

                            print ("\nLos datos de tu reservacion son los siguientes")

                            print ("\n--------------------------------------------------------------")
                            print ("Clave       Nombre        Asistentes        Fecha        Turno")
                            print (" ",evento[0][0],"      ",evento[0][1],"        ",evento[0][2],"      ",evento[0][3],"      ",evento[0][4])
                            print ("--------------------------------------------------------------")

                            print ("\n¿Estas segur@ de que deseas eliminar tu reservacion?")
                            print ("\nUna vez eliminada la reservacion no podra recuperarse")
                            print ('\n[1] "SI" [2] "NO"')

                            eliminar = int (input ("\nIngresa tu respuesta : "))

                            if eliminar == 1 :

                                fecha_empaquetar = fecha[0][0]
                                invertido = (list(reversed(fecha_empaquetar)))
                                fecha_invertida = (invertido[1] + invertido[0] + "/" + invertido[4] + invertido[3] + "/" + invertido[9] + invertido[8] + invertido[7] + invertido[6])
                                fecha_invertida_convertida = datetime.datetime.strptime (fecha_invertida,'%d/%m/%Y').date()

                                if fecha_invertida_convertida <= fecha_hoy :
                                    print ("\n-------------------------------------------------------------------------------")
                                    print ("Su reservacion no puede ser eliminada / No cuenta con tres dias de anticipacion")
                                    print ("-------------------------------------------------------------------------------")

                                    regreso = input ('\nEscribe "0" para regresar al menu : ')

                                elif fecha_invertida_convertida == fecha_hoy + datetime.timedelta(days=+dia) :
                                    print ("\n-------------------------------------------------------------------------------")
                                    print ("Su reservacion no puede ser eliminada / No cuenta con tres dias de anticipacion")
                                    print ("-------------------------------------------------------------------------------")

                                    regreso = input ('\nEscribe "0" para regresar al menu : ')

                                elif fecha_invertida_convertida <= fecha_hoy + datetime.timedelta(days=+dia_dos) :

                                    print ("\n-------------------------------------------------------------------------------")
                                    print ("Su reservacion no puede ser eliminada / No cuenta con tres dias de anticipacion")
                                    print ("-------------------------------------------------------------------------------")

                                    regreso = input ('\nEscribe "0" para regresar al menu : ')

                                elif fecha_invertida_convertida >= fecha_hoy + datetime.timedelta(days=+dia_tres) :

                                    try:
                                        with sqlite3.connect("Evidencia.db") as conn:
                                            mi_cursor = conn.cursor()
                                            valores = {"Clave_evento":nuevo_acceso}

                                            mi_cursor.execute("DELETE FROM Eventos_registrados WHERE Clave_evento = :Clave_evento", valores)

                                    except Error as e:
                                        print (e)

                                    except:
                                        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")

                                    finally:
                                        conn.close()

                                    print ("\n-------------------------------------------")
                                    print ("Su reservacion se ha eliminado exitosamente")
                                    print ("-------------------------------------------")

                                    regreso = input ('\nEscribe "0" para regresar al menu : ')

                            elif eliminar == 2 :
                                print ("\n----------------------------")
                                print ("Su reservacion no se elimino")
                                print ("----------------------------")

                                regreso = input ('\nEscribe "0" para regresar al menu : ')

                            else :
                                print ("\n------------------------------------------------------------------")
                                print ("La opcion que escribio no es valida / Por favor intentalo de nuevo")
                                print ("------------------------------------------------------------------")

                                regreso = input ('\nEscribe "0" para regresar al menu : ')

                        else :
                            print ("\n--------------------")
                            print ("Evento no encontrado")
                            print ("--------------------")

                            regreso = input ('\nEscribe "0" para regresar al menu : ')

                except Error as e:
                    print (e)

                except:
                    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")

                finally:
                    conn.close()

            # Regresar al menu principal

            else :
                break

    # Menu opcion 2

    elif menu_principal == 2 :

        while True :
            print ("\n-- REVISAR LOS REPORTES --")

            print ("\nMENU")
            print ('\n[1] Consultar las reservaciones en "pantalla"')
            print ('\n[2] Consultar las reservaciones en "excel"')
            print ("\n[3] Regresar al menu principal")

            sub_menu = int (input ("\nPor favor escribe el numero del menu al que desea acceder : "))

            # Reservaciones en pantalla

            if sub_menu == 1 :
                print ('\n-- CONSULTAR LAS RESERVACIONES EN "PANTALLA" --')

                try:
                    with sqlite3.connect("Evidencia.db") as conn:
                        mi_cursor = conn.cursor()
                        mi_cursor.execute("CREATE TABLE IF NOT EXISTS Eventos_registrados (Clave_evento INTEGER PRIMARY KEY, Nombre_cliente TEXT NOT NULL, Nombre_sala TEXT NOT NULL, Asistentes INTEGER, Fecha timestamp, Turno INTEGER);")

                except Error as e:
                    print (e)

                except:
                    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")

                print ('\nEscribe la fecha con el siguiente formato "09/01/2004"')
                consulta = input("\nEscribe la fecha en la que quiere consultar las reservaciones : ")
                consulta = datetime.datetime.strptime (consulta,'%d/%m/%Y').date()

                try:
                    with sqlite3.connect("Evidencia.db") as conn:
                        mi_cursor = conn.cursor()
                        valores = {"Fecha":consulta}

                        mi_cursor.execute("SELECT Clave_evento, Nombre_cliente, Nombre_sala, Turno FROM Eventos_registrados WHERE Fecha = :Fecha", valores)
                        reporte_pantalla = mi_cursor.fetchall()

                        print ("\n------------------------------------------------------------------")
                        print (f"                    Reporte del dia {consulta}")
                        print ("------------------------------------------------------------------")
                        print ("Clave de la sala       Cliente       Nombre del evento       Turno")
                        print ("------------------------------------------------------------------")

                        for clave,sala,nombre,turno in reporte_pantalla :
                            print ("      ",clave,"           ",sala,"       ",nombre,"            ",turno)
                            print ("------------------------------------------------------------------")

                except Error as e:
                    print (e)

                except:
                    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")

                finally:
                    conn.close()

                    regreso = input ('\nEscribe "0" para regresar al menu : ')

            # Reservaciones en excel

            elif sub_menu == 2 :
                print ('\n-- CONSULTAR LAS RESERVACIONES EN "EXCEL" --')

                try:
                    with sqlite3.connect("Evidencia.db") as conn:
                        mi_cursor = conn.cursor()
                        mi_cursor.execute("CREATE TABLE IF NOT EXISTS Eventos_registrados (Clave_evento INTEGER PRIMARY KEY, Nombre_cliente TEXT NOT NULL, Nombre_sala TEXT NOT NULL, Asistentes INTEGER, Fecha timestamp, Turno INTEGER);")

                except Error as e:
                    print (e)

                except:
                    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")

                print ('\nEscribe la fecha con el siguiente formato "09/01/2004"')
                consulta_tres = input("\nEscribe la fecha en la que quiere consultar las reservaciones : ")
                consulta_tres = datetime.datetime.strptime (consulta_tres,'%d/%m/%Y').date()

                try:
                    with sqlite3.connect("Evidencia.db") as conn:
                        mi_cursor = conn.cursor()
                        valores = {"Fecha":consulta_tres}

                        mi_cursor.execute("SELECT Clave_evento, Nombre_cliente, Nombre_sala, Turno FROM Eventos_registrados WHERE Fecha = :Fecha", valores)
                        reporte_excel = mi_cursor.fetchall()

                        valores = {"Fecha":consulta_tres}

                        libro = openpyxl.Workbook()
                        hoja = libro["Sheet"]
                        hoja.title = "Primera"

                        hoja["C2"].value = "Reporte para el dia : "
                        hoja["D2"].value = consulta_tres
                        hoja["B4"].value = "Clave de la sala"
                        hoja["C4"].value = "Cliente"
                        hoja["D4"].value = "Nombre del evento"
                        hoja["E4"].value = "Turno"

                        for clave,sala,nombre,turno in reporte_excel :
                            hoja.cell(row=clave + 5,column=2).value = clave
                            hoja.cell(row=clave + 5,column=3).value = sala
                            hoja.cell(row=clave + 5,column=4).value = nombre
                            hoja.cell(row=clave + 5,column=5).value = turno

                        libro.save("MiExcelDesdePython.xlsx")

                except Error as e:
                    print (e)

                except:
                    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")

                finally:
                    conn.close()

                    print ("\n-----------------------------------------------------------")
                    print ("Reporte creado exitosamente / Por favor revisa el documento")
                    print ("-----------------------------------------------------------")

                    regreso = input ('\nEscribe "0" para regresar al menu : ')

            # Regresar al menu principal

            else :
                break

    # Menu opcion 3

    elif menu_principal == 3 :
        print ("\n-- REGISTRAR UNA SALA --")

        try:
            with sqlite3.connect("Evidencia.db") as conn:
                mi_cursor = conn.cursor()
                mi_cursor.execute("CREATE TABLE IF NOT EXISTS Clientes_registrados (Clave_cliente INTEGER PRIMARY KEY, Nombre_cliente TEXT NOT NULL);")

        except Error as e:
            print (e)

        except:
            print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")

        print ("\nPara registrar una sala es necesario ser cliente registrado")
        acceso = int (input("\nPor favor ingresa tu clave como cliente : "))

        # Registrar sala

        # Validacion cliente

        try:
            with sqlite3.connect("Evidencia.db") as conn:
                mi_cursor = conn.cursor()
                valores = {"Clave_cliente": acceso}
                mi_cursor.execute("SELECT Nombre_cliente FROM Clientes_registrados WHERE Clave_cliente = :Clave_cliente", valores)
                nombre_cliente = mi_cursor.fetchall()

                if nombre_cliente :
                    print ("\n------------------------------------------------")
                    print ("Cliente encontrado / Bienvenid@" , nombre_cliente[0][0] )
                    print ("------------------------------------------------")

                    # Nombre del evento

                    while True :
                        sala = input ("\nEscribe el nombre del evento para registrar tu sala : ")

                        if sala == "" :
                            print ("\n------------------------------")
                            print ("El nombre no puede ser omitido")
                            print ("------------------------------")

                        else :
                            print ("\n-----------------")
                            print ("Nombre registrado")
                            print ("-----------------")
                            break

                    # Asistentes

                    while True :
                        try :
                            asistentes = int (input("\nEscribe la cantidad de asistentes para tu evento : "))

                        except Exception :
                            print ("\n-----------------------------")
                            print ("El valor no puede ser omitido")
                            print ("-----------------------------")

                        else :

                            if asistentes == 0 :
                                print ("\n-------------------------------------------------")
                                print ('La cantidad de asistentes debe de ser mayor a "0"')
                                print ("-------------------------------------------------")

                            elif asistentes > 0 :
                                print ("\n----------------------")
                                print ("Asistentes registrados")
                                print ("----------------------")
                                break

                    # Fecha

                    while True :
                        print ("\nEscribre la fecha para tu evento con el siguiente formato (Dia/Mes/Año)")
                        print ('\nEjemplo : "09/01/2004"')
                        print ("\nLa fecha debe de ser con dos dias de anticipacion")

                        fecha = input ("\nEscribe la fecha para tu evento : ")
                        fecha = datetime.datetime.strptime (fecha,'%d/%m/%Y').date()

                        if fecha <= fecha_hoy :
                            print ("\n-------------------------------------------------------------------")
                            print ("Fecha no valida / La fecha debe de ser con dos dias de anticipacion")
                            print ("-------------------------------------------------------------------")

                        elif fecha == fecha_hoy + datetime.timedelta(days=+dia) :
                            print ("\n-------------------------------------------------------------------")
                            print ("Fecha no valida / La fecha debe de ser con dos dias de anticipacion")
                            print ("-------------------------------------------------------------------")

                        elif fecha >= fecha_hoy + datetime.timedelta(days=+dia_dos) :
                            print ("\n----------------")
                            print ("Fecha disponible")
                            print ("----------------")
                            break

                    try:
                        with sqlite3.connect("Evidencia.db") as conn:
                            mi_cursor = conn.cursor()
                            mi_cursor.execute("CREATE TABLE IF NOT EXISTS Eventos_registrados (Clave_evento INTEGER PRIMARY KEY, Nombre_cliente TEXT NOT NULL, Nombre_sala TEXT NOT NULL, Asistentes INTEGER, Fecha timestamp, Turno INTEGER);")

                    except Error as e:
                        print (e)

                    except:
                        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")

                    try:
                        with sqlite3.connect("Evidencia.db") as conn:
                            mi_cursor = conn.cursor()

                            criterios = {"Fecha":fecha}
                            mi_cursor.execute("SELECT Turno FROM Eventos_registrados WHERE DATE(Fecha) = :Fecha;", criterios)
                            turnos_ocupados = mi_cursor.fetchall()
                            turnos_ocupados = list (turnos_ocupados)
                            turnos_ocupados.append(("Python",0))
                            turnos_ocupados.append(("Python",0))
                            turnos_ocupados.append(("Python",0))

                    except Error as e:
                        print (e)

                    except:
                        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")

                    finally:
                        conn.close()

                    # Turno

                    while True :

                        print ("\nHorarios disponibles")
                        print ("\n[1] Mañana")
                        print ("\n[2] Tarde")
                        print ("\n[3] Noche")

                        turno = int (input ("\nIngresa el turno en el que deseas registrar tu sala : "))

                        if turno >= 4 :
                            print ("\n------------------------------")
                            print ("El turno que ingreso no existe")
                            print ("------------------------------")

                        elif turno in turnos_ocupados[0] or turno in turnos_ocupados[1] or turno in turnos_ocupados[2] :
                            print ("\n------------------------------------------------------------------")
                            print ("Turno no disponible / Ya existe un evento registrado en este turno")
                            print ("------------------------------------------------------------------")

                        else :
                            print ("\n----------------")
                            print ("Turno disponible")
                            print ("----------------")
                            turnos_ocupados.remove(("Python",0))
                            turnos_ocupados.remove(("Python",0))
                            turnos_ocupados.remove(("Python",0))
                            break

                    # Base de datos eventos

                    try:
                        with sqlite3.connect("Evidencia.db") as conn:
                            mi_cursor = conn.cursor()
                            valores = {"Nombre_cliente":nombre_cliente[0][0],"Nombre_sala":sala,"Asistentes":asistentes,"Fecha":fecha,"Turno":turno}
                            mi_cursor.execute("INSERT INTO Eventos_registrados (Nombre_cliente, Nombre_sala, Asistentes, Fecha, Turno) VALUES (:Nombre_cliente,:Nombre_sala,:Asistentes,:Fecha,:Turno)", valores)
                            clave_evento = mi_cursor.lastrowid

                    except Error as e:
                        print (e)

                    except:
                        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")

                    finally:
                        conn.close()

                    # Registro terminado

                    print ("-------------------------------------")
                    print ("Tu sala se ha registrado exitosamente")
                    print ("La clave de tu sala es : ",clave_evento)
                    print ("-------------------------------------")

                    datos = input('\nIngresa "0" para ver los datos de tu reservacion : ')

                    print ("\nLos datos de tu reservacion son los siguientes")
                    print ("\n-------------------------------------------------------------------")
                    print("Clave","       ","Nombre","       ","Asistentes","       ","Fecha","       ","Turno")
                    print (f"  {clave_evento}  \t     {sala} \t        {asistentes}          {fecha}    \t{turno} ")
                    print ("-------------------------------------------------------------------")

                    # Base de datos eventos Fecha - Turno - Universal

                    try:
                        with sqlite3.connect("Evidencia.db") as conn:
                            mi_cursor = conn.cursor()
                            mi_cursor.execute("CREATE TABLE IF NOT EXISTS Fecha_turno (Fecha timestamp, Turno INTEGER);")

                    except Error as e:
                        print (e)

                    except:
                        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")

                    try:
                        with sqlite3.connect("Evidencia.db") as conn:
                            mi_cursor = conn.cursor()
                            valores = {"Fecha":fecha,"Turno":1}
                            mi_cursor.execute("INSERT INTO Fecha_turno (Fecha, Turno) VALUES (:Fecha,:Turno)", valores)

                            valores = {"Fecha":fecha,"Turno":2}
                            mi_cursor.execute("INSERT INTO Fecha_turno (Fecha, Turno) VALUES (:Fecha,:Turno)", valores)

                            valores = {"Fecha":fecha,"Turno":3}
                            mi_cursor.execute("INSERT INTO Fecha_turno (Fecha, Turno) VALUES (:Fecha,:Turno)", valores)

                    except Error as e:
                        print (e)

                    except:
                        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")

                    finally:
                        conn.close()

                    regreso = input ('\nIngresa "0" para regresar al menu principal : ')

                # Validacion cliente

                else :
                    print ("\n---------------------")
                    print ("Cliente no encontrado")
                    print ("---------------------")

                    regreso = input ('\nEscribe "0" para regresar al menu principal : ')

        except Error as e:
            print (e)

        except:
            print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")

        finally:
            conn.close()

    # Menu opcion 4

    elif menu_principal == 4 :
        print ("\n-- REGISTRARSE COMO NUEVO CLIENTE --")

        while True :
            cliente_nuevo = input ("\nPor favor escribe tu nombre completo : ")

            if cliente_nuevo == "" :
                print ("\n------------------------------")
                print ("El nombre no puede ser omitido")
                print ("------------------------------")

            # Nuevo cliente

            # Base de datos

            else:
                try:
                    with sqlite3.connect("Evidencia.db") as conn:
                        mi_cursor = conn.cursor()
                        mi_cursor.execute("CREATE TABLE IF NOT EXISTS Clientes_registrados (Clave_cliente INTEGER PRIMARY KEY, Nombre_cliente TEXT NOT NULL);")

                except Error as e:
                    print (e)

                except:
                    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")

                try:
                    with sqlite3.connect("Evidencia.db") as conn:
                        mi_cursor = conn.cursor()
                        valores = {"Nombre_cliente": cliente_nuevo}
                        mi_cursor.execute("INSERT INTO Clientes_registrados (Nombre_cliente) VALUES (:Nombre_cliente)", valores)

                        print ("\n---------------------------------------------------------")
                        print (f"Felicidades {cliente_nuevo} tu registro concluyo exitosamente")
                        print (f"\nTu clave es : {mi_cursor.lastrowid}")
                        print ("---------------------------------------------------------")

                except Error as e:
                    print (e)

                except:
                    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")

                finally:
                    conn.close()
                    break

        regreso = input ('\nEscribe "0" para regresar al menu principal : ')

    # Menu opcion 5

    elif menu_principal == 5 :
        print ("\n-- SALIR DEL PROGRAMA --")

        # Salida

        print ("\n-------------------------")
        print ("El programa ha finalizado")
        print ("-------------------------")
        break

    # Menu dato no valido

    else :
        print ("\n------------------------------------------------------------------")
        print ("La opcion que escribio no es valida / Por favor intentalo de nuevo")
        print ("------------------------------------------------------------------")

        regreso = input ('\nIngresa "0" para regresar al menu principal : ')